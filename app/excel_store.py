from __future__ import annotations
import logging
import threading
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook

from app.models import BulkTransactionResponse, Currency, Transaction, TransactionCreate, TransactionType, TransactionUpdate

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
EXCEL_PATH = DATA_DIR / "expenses.xlsx"

logger = logging.getLogger("uvicorn.error")
TRANSACTION_HEADERS = [
    "id",
    "date",
    "type",
    "category",
    "amount",
    "currency",
    "description",
    "created_at",
]
CATEGORY_HEADERS = ["name", "type", "keywords"]

DEFAULT_CATEGORIES = [
    ("Food", "expense", ""),
    ("Transport", "expense", ""),
    ("Shopping", "expense", ""),
    ("Bills", "expense", ""),
    ("Entertainment", "expense", ""),
    ("Healthcare", "expense", ""),
    ("Other Expense", "expense", ""),
    ("Salary", "income", ""),
    ("Freelance", "income", ""),
    ("Investment", "income", ""),
    ("Other Income", "income", ""),
]

_lock = threading.Lock()


def _ensure_workbook() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if EXCEL_PATH.exists():
        return

    wb = Workbook()
    tx_sheet = wb.active
    tx_sheet.title = "Transactions"
    tx_sheet.append(TRANSACTION_HEADERS)

    cat_sheet = wb.create_sheet("Categories")
    cat_sheet.append(CATEGORY_HEADERS)
    for name, cat_type, keywords in DEFAULT_CATEGORIES:
        cat_sheet.append([name, cat_type, keywords])

    wb.save(EXCEL_PATH)


def _load_transactions_sheet():
    _ensure_workbook()
    wb = load_workbook(EXCEL_PATH)
    return wb, wb["Transactions"]


def _row_to_transaction(row: list) -> Transaction:
    return Transaction(
        id=int(row[0]),
        date=date.fromisoformat(str(row[1])),
        type=TransactionType(str(row[2])),
        category=str(row[3]),
        amount=float(row[4]),
        currency=Currency(str(row[5])),
        description=str(row[6] or ""),
    )


def _next_id(sheet) -> int:
    max_id = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            max_id = max(max_id, int(row[0]))
    return max_id + 1


def list_transactions(
    year: Optional[int] = None,
    month: Optional[int] = None,
    day: Optional[int] = None,
) -> list[Transaction]:
    with _lock:
        wb, sheet = _load_transactions_sheet()
        transactions: list[Transaction] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            tx = _row_to_transaction(list(row))
            if year and tx.date.year != year:
                continue
            if month and tx.date.month != month:
                continue
            if day and tx.date.day != day:
                continue
            transactions.append(tx)
        wb.close()
        transactions.sort(key=lambda t: (t.date, t.id))
        return transactions


def get_transaction(tx_id: int) -> Optional[Transaction]:
    with _lock:
        wb, sheet = _load_transactions_sheet()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if int(row[0]) == tx_id:
                tx = _row_to_transaction(list(row))
                wb.close()
                return tx
        wb.close()
        return None


def create_transaction(payload: TransactionCreate) -> Transaction:
    with _lock:
        wb, sheet = _load_transactions_sheet()
        tx_id = _next_id(sheet)
        now = datetime.now().isoformat(timespec="seconds")
        sheet.append(
            [
                tx_id,
                payload.date.isoformat(),
                payload.type.value,
                payload.category,
                payload.amount,
                payload.currency.value,
                payload.description,
                now,
            ]
        )
        wb.save(EXCEL_PATH)
        wb.close()
        return Transaction(id=tx_id, **payload.model_dump())


def create_bulk_transactions(payloads: list[TransactionCreate]) -> BulkTransactionResponse:
    with _lock:
        wb, sheet = _load_transactions_sheet()
        tx_id = _next_id(sheet)
        now = datetime.now().isoformat(timespec="seconds")
        created: list[Transaction] = []
        for payload in payloads:
            logger.info(f"Creating transaction: txId: {tx_id}, date:{payload.date}, dateIso: {payload.date.isoformat()}, amount: {payload.amount}")
            sheet.append(
                [
                    tx_id,
                    payload.date.isoformat(),
                    payload.type.value,
                    payload.category,
                    payload.amount,
                    payload.currency.value,
                    payload.description,
                    now,
                ]
            )
            created.append(Transaction(id=tx_id, **payload.model_dump()))
            tx_id += 1
        wb.save(EXCEL_PATH)
        wb.close()
        return BulkTransactionResponse(count=len(created), transactions=created)


def update_transaction(tx_id: int, payload: TransactionUpdate) -> Optional[Transaction]:
    with _lock:
        wb, sheet = _load_transactions_sheet()
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if row[0].value is None:
                continue
            if int(row[0].value) != tx_id:
                continue

            current = _row_to_transaction([cell.value for cell in row])
            updates = payload.model_dump(exclude_unset=True)
            data = current.model_dump()
            data.update(updates)
            if "type" in updates and updates["type"] is not None:
                data["type"] = updates["type"]
            if "currency" in updates and updates["currency"] is not None:
                data["currency"] = updates["currency"]

            sheet.cell(row=idx, column=2, value=data["date"].isoformat())
            sheet.cell(row=idx, column=3, value=data["type"].value)
            sheet.cell(row=idx, column=4, value=data["category"])
            sheet.cell(row=idx, column=5, value=data["amount"])
            sheet.cell(row=idx, column=6, value=data["currency"].value)
            sheet.cell(row=idx, column=7, value=data["description"])

            wb.save(EXCEL_PATH)
            wb.close()
            return Transaction(**data)

        wb.close()
        return None


def delete_transaction(tx_id: int) -> bool:
    with _lock:
        wb, sheet = _load_transactions_sheet()
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if row[0].value is None:
                continue
            if int(row[0].value) == tx_id:
                sheet.delete_rows(idx)
                wb.save(EXCEL_PATH)
                wb.close()
                return True
        wb.close()
        return False


def list_categories(tx_type: Optional[str] = None) -> list[dict[str, str]]:
    with _lock:
        _ensure_workbook()
        wb = load_workbook(EXCEL_PATH)
        sheet = wb["Categories"]
        categories: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            name, cat_type = str(row[0]), str(row[1])
            keywords = str(row[2] or "") if len(row) > 2 else ""
            if tx_type and cat_type != tx_type:
                continue
            categories.append({"name": name, "type": cat_type, "keywords": keywords})
        wb.close()
        return categories


def create_category(name: str, cat_type: str, keywords: str = "") -> dict[str, str]:
    with _lock:
        _ensure_workbook()
        wb = load_workbook(EXCEL_PATH)
        sheet = wb["Categories"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if str(row[0]) == name and str(row[1]) == cat_type:
                wb.close()
                return {"error": "Category already exists"}
        sheet.append([name, cat_type, keywords])
        wb.save(EXCEL_PATH)
        wb.close()
        return {"name": name, "type": cat_type, "keywords": keywords}


def update_category(old_name: str, new_name: str, cat_type: str, keywords: Optional[str] = None) -> dict[str, str]:
    with _lock:
        _ensure_workbook()
        wb = load_workbook(EXCEL_PATH)
        sheet = wb["Categories"]
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if row[0].value is None:
                continue
            if str(row[0].value) == old_name and str(row[1].value) == cat_type:
                sheet.cell(row=idx, column=1, value=new_name)
                if keywords is not None:
                    sheet.cell(row=idx, column=3, value=keywords)
                else:
                    keywords = str(row[2].value or "") if len(row) > 2 else ""
                wb.save(EXCEL_PATH)
                wb.close()
                return {"name": new_name, "type": cat_type, "keywords": keywords}
        wb.close()
        return {"error": "Category not found"}


def delete_category(name: str, cat_type: str) -> bool:
    with _lock:
        _ensure_workbook()
        wb = load_workbook(EXCEL_PATH)
        sheet = wb["Categories"]
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if row[0].value is None:
                continue
            if str(row[0].value) == name and str(row[1].value) == cat_type:
                sheet.delete_rows(idx)
                wb.save(EXCEL_PATH)
                wb.close()
                return True
        wb.close()
        return False


def list_currencies() -> list[str]:
    return [c.value for c in Currency]
